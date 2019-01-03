#! /usr/bin/python

def load(src: str, idx: int):
  from openpyxl import load_workbook
  wb = load_workbook(src, read_only = True)
  wx = wb[wb.sheetnames[0]]
  model = []
  for row in wx.rows:
    line = []
    for cell in row:
      if cell.value and cell.value.find('|') != -1:
        print("Cell cannot contains '|' which is a cell seperator")
        exit(-1)
      line.append(cell.value if cell.value else '')
    model.append(line)
  return model

def save(model, dst: str):
  maxwidths = [0] * len(model[0])
  for i in range(len(model)):
    for j in range(len(model[i])):
      celllen = max(map(lambda x: len(x), model[i][j].split('\n')))
      celllen = (celllen + 2) if celllen > 0 else 1
      if celllen > maxwidths[j]:
        maxwidths[j] = celllen
  linesplitor = '+%s+' % '+'.join(map(lambda x: '-' * x if x > 0 else '-', maxwidths))
  with open(dst, 'w') as out:
    out.write(linesplitor + '\n')
    for i in range(len(model)):
      maxline = 0
      for j in range(len(model[i])):
        cell = model[i][j]
        linecnt = len(cell.split('\n'))
        if linecnt > maxline:
          maxline = linecnt
      row = []
      for j in range(len(model[i])):
        cell = model[i][j]
        lines = cell.split('\n')
        if len(lines) < maxline:
          row.append([''] * (maxline - len(lines)) + lines)
        else:
          row.append(lines)
      for k in range(maxline):
        ln = []
        for j in range(len(model[i])):
          ln.append(' ' + row[j][k].ljust(maxwidths[j] - 1))
        out.write('|%s|\n' % '|'.join(ln))
      out.write(linesplitor + '\n')

def main(src: str, dst: str, idx: int):
  model = load(src, idx)
  save(model, dst)

if __name__ == '__main__':
  import argparse
  import sys
  parser = argparse.ArgumentParser()
  parser.add_argument("src", help="The excel file")
  parser.add_argument("dst", help="The table file")
  parser.add_argument("--sheet-index", dest="idx", type=int, default=0, help="The index of sheet")
  args = parser.parse_args()
  main(args.src, args.dst, args.idx)
